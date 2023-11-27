from openpyxl import load_workbook
### I HATE python SO much for making me do this utter bs to import from a parent directory
import os
import sys
current_directory = os.path.dirname(os.path.realpath(__file__))
project_directory = os.path.dirname(os.path.dirname(current_directory))
sys.path.append(project_directory)

from Stuff.Server import Server, is_valid_ip
from Stuff.Excel_helpers import is_cell_green, is_cell_orange, get_cell_fill_color

## These guys are the general assumption I make where the relevant data should be.
## Changing them widens or tightens the search area the count function runs on, and effect the speed and accuracy of the server discovery
MAX_ROW_RANGE = 28
MIN_ROW_RANGE = 7

MAX_COL_RANGE = 5
MIN_COL_RANGE = 1
# Change if the search area is too small its missing servers, or too big its grabbing other non relevant servers



def main(sheet_name, the_servers: list, debug_mode: bool):
    
    implementation_type: str = ""
    match count_C_app_servers(sheet_name, the_servers, get_C_app_implementation_type(sheet_name)):
        case 1:
            implementation_type = "Single Server"
        case 2:
            implementation_type = "Single Server and Reporting"
        case 4:
            implementation_type = "HA"
        case 5:
            implementation_type = "HA with 1 Reporting"
        case 6:
            implementation_type = "HA with 2 Reporting"
        case _:
            raise Exception("Could not determine proper deployment \nPlease highlight correct servers Green (\"Good input\" on excel style picker) and optionally old servers as Red (\"Bad\" input) \nAborting")
    
    print("Read as", implementation_type)
    
    if(is_valid_deployment(the_servers, implementation_type)):
        print("Verified as", implementation_type)

  
    #DEBUG#############################
    if debug_mode:
        for server in the_servers:
            print(server)
    #DEBUG#############################

 














  



def get_C_app_implementation_type(sheet_name) -> str:
    """
    Grabs the Implementation type cell
    This is important because I use it to validate the environment makes sense as in its a supported type of deployment,
    and helps filter out possible ambiguous servers that should not be looked at.

    Args:
        sheet_name (Openpyxl.Workbook.Sheet): Excel Workbook sheet

    Raises:
        ValueError: Throws a value error if we couldn't determine the implementation_type. So it must stop.

    Returns:
        str: Returns the implementation_type str var
    """
    implementation_type: str = ""
    implementation_type = str(sheet_name.cell(row = 3, column = 2).value)
    
    match implementation_type:
        case "Migration":
            print("Determined Implementation type as:", implementation_type)
            print("**************************************************************")
            print("Please Highlight New Servers Green aka \"Good input in excel\"")
            print("**************************************************************")
        case "Upgrade":
            print("Determined Implementation type as:", implementation_type)
            print("**************************************************************")
            print("Please Highlight New Servers Green aka \"Good input in excel\"")
            print("**************************************************************")
        case "New":
            print("Determined Implementation type as:", implementation_type)
        case _:
            raise ValueError("Unable to determine Implementation type of this deployment\nEnsure that the drop down box in cell B3 is set.")
    return implementation_type
    

 
def count_C_app_servers(sheet_name, list_of_servers, implementation_type) -> int:
    """
    Counts the valid recongized servers and builds the Server object list

    Args:
        sheet_name (Openpyxl.Workbook.Sheet): Excel Workbook sheet
        list_of_servers (list): Empty list which will be a list of "Server" objects
        implementation_type (str): The string corresponding to the implementation type

    Returns:
        int: the number of servers counted
    """
    count = 0
    if implementation_type == "Migration" or "Upgrade":
        # check for green cells in counting because migrations & upgrades have the old and new servers which would lead to ambiguous counts
        for row in range(MIN_ROW_RANGE, MAX_ROW_RANGE):
            for column in range(MIN_COL_RANGE, MAX_COL_RANGE):
                # If the cell is empty move along
                if sheet_name.cell(row, column).value != None:
                    if is_valid_ip(str(sheet_name.cell(row, column).value)) and is_cell_green(sheet_name, row, column):
                        # Debug
                        # print(str(sheet_name.cell(row, column).value), "Coord:", row, column)
                        # Check to see if there's a provided username/password. If not assign default logins
                        if str(sheet_name.cell(row, column+5).value == "None"):
                            username = "user"
                        else:
                            username = str(sheet_name.cell(row, column+5).value)
                        if str(sheet_name.cell(row, column+6).value == "None"):
                            password = "password@123"
                        else:
                            password= str(sheet_name.cell(row, column+6).value)
                        list_of_servers.append(Server(sheet_name.cell(row, column-1).value, sheet_name.cell(row, column).value, sheet_name.cell(row, column+2).value, sheet_name.cell(row, column+1).value, row, column, username, password))
                        count += 1       
    else:
        for row in range(MIN_ROW_RANGE, MAX_ROW_RANGE):
            for column in range(MIN_COL_RANGE, MAX_COL_RANGE):
                # If the cell is empty move along
                if sheet_name.cell(row, column).value != None:
                    if is_valid_ip(str(sheet_name.cell(row, column).value)) and (is_cell_green(sheet_name, row, column) or is_cell_orange(sheet_name, row, column) ):
                        # Debug
                        # print(str(sheet_name.cell(row, column).value), "Coord:", row, column)
                        # Check to see if there's a provided username/password. If not assign default logins
                        if str(sheet_name.cell(row, column+5).value == "None"):
                            username = "user"
                        else:
                            username = str(sheet_name.cell(row, column+5).value)
                        if str(sheet_name.cell(row, column+6).value == "None"):
                            password = "password@123"
                        else:
                            password= str(sheet_name.cell(row, column+6).value)
                        list_of_servers.append(Server(sheet_name.cell(row, column-1).value, sheet_name.cell(row, column).value, sheet_name.cell(row, column+2).value, sheet_name.cell(row, column+1).value, row, column, username, password))
                        count += 1
    #TODO double check the logic here^
    return count


def is_valid_deployment(list_of_servers: list, implementation_type: str) -> bool:
    """
    This function is messy but it makes sense to me and I have so much more code to write here so I'm not redoing this logic.
    Basically want to do one last thorough sanity check that everything is as expected and the program from here on knows exactly what data it has.
    Also now it assigns a port to each server. I used to do this in a different function later on but might as well do it here and save running through the loop again later.

    Args:
        list_of_servers (list): This is the list of Server objects. Its used to count the servers and access their data now
        implementation_type (str): The string returned from earlier match statment which used the count_C_app_servers function to deduce

    Returns:
        bool: Returns a bool. The program needs this because without it, we're just taking guesses and what server is what. That's not bueono because we do different commands for different server types
    """
    match implementation_type:
        case "Single Server":
            if len(list_of_servers) == 1:
                list_of_servers[0].port = 6001
                return True
        case "Single Server and Reporting":
            if len(list_of_servers) == 2:
                list_of_servers[0].port = 6001
                list_of_servers[1].port = 6002
                return True
        case "HA":
            if len(list_of_servers) == 4:
                flag1: bool = False
                flag2: bool = False
                app_server_count: int = 0
                for server in list_of_servers:
                    match server._ha_type:
                        case "Application Server":
                            app_server_count += 1
                            if app_server_count == 1:
                                server.port = 6001
                            else:
                                server.port = 7001
                        case "Master Database":
                            flag1 = True
                            server.port = 6002
                        case "Standby Database":
                            flag2 = True
                            server.port = 7002
                if flag1 and flag2 and (app_server_count == 2):
                    return True
        case "HA with 1 Reporting":
            if len(list_of_servers) == 5:
                flag1: bool = False
                flag2: bool = False
                flag3: bool = False
                app_server_count: int = 0
                for server in list_of_servers:
                    match server._ha_type:
                        case "Application Server":
                            # The logic here is annoying and not bullet proof but its going to be over engineered and very long and at worse case if this goes wrong it just makes for a mismatched port forward between a DC later on
                            # Basically we assume the first app server we parse is the primary. My old program used the subnet of the first server in the array but i dont feel like doing that right now and this works close enough
                            #:6000's are DC1 :7000s are DC2, this really is only used in the port forwards and the port is a arbitrary one I go with and use in the SuperPuTTY profiles so this is all a waste of time explaining
                            app_server_count += 1
                            if app_server_count == 1:
                                server.port = 6001
                            else:
                                server.port = 7001
                        case "Master Database":
                            flag1 = True
                            server.port = 6002
                        case "Standby Database":
                            flag2 = True
                            server.port = 7002
                        case "Primary Reporting":
                            flag3 = True
                            server.port = 6003
                if flag1 and flag2 and flag3 and (app_server_count == 2):
                    return True
        case "HA with 2 Reporting":
            if len(list_of_servers) == 6:
                flag1: bool = False
                flag2: bool = False
                flag3: bool = False
                flag4: bool = False
                app_server_count: int = 0
                for server in list_of_servers:
                    match server._ha_type:
                        case "Application Server":
                            app_server_count += 1
                            if app_server_count == 1:
                                server.port = 6001
                            else:
                                server.port = 7001
                        case "Master Database":
                            flag1 = True
                            server.port = 6002
                        case "Standby Database":
                            flag2 = True
                            server.port = 7002
                        case "Primary Reporting":
                            flag3 = True
                            server.port = 6003
                        case "DR Reporting":
                            flag4 = True
                            server.port = 7003
                if flag1 and flag2 and flag3 and flag4 and (app_server_count == 2):
                    return True
        case _:
            return False























# Testing parameters
if __name__ =="__main__":
    workbook = load_workbook(filename="/home/dean/test.xlsx", read_only=True)
    sheet_name = workbook["C"]
    test_list = []
    debug_mode = True
    main(sheet_name, test_list, debug_mode)