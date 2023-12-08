import paramiko
from openpyxl import load_workbook, worksheet
import os
import sys

# Again this is so wack that I have to do this.
current_directory = os.path.dirname(os.path.realpath(__file__))
project_directory = os.path.dirname(current_directory)
sys.path.append(project_directory)

from Stuff.Server import Server

def generate_hosts_file(list_of_servers: list[Server], debug_mode: bool) -> None:
    """ This creates the host file and overwrites old ones if found

    Args:
        list_of_servers (list[Server]): The list of server objects 
        debug_mode (bool): whether to print debugging lines or not
    """
    # Check if the hosts file exists, and delete it, if it does, because this would append to a old one if not done
    file_path="hosts"
    if os.path.exists(file_path):
        if debug_mode:
            print("Old hosts file exists, deleting and making a new one")
            os.remove(file_path)
    
    with open(file_path, 'w') as hosts_file:
        # throw in the localhost lines manually
        hosts_file.write("127.0.0.1\tlocalhost localhost.localdomain localhost6 localhost6.localdomain6\n")
        hosts_file.write("::1     \tlocalhost localhost.localdomain localhost6 localhost6.localdomain6\n")
        for server in list_of_servers:
            hosts_file.write(f"{server.ip_address}\t{server.fqdn}\t{server._hostname}\n")
        if debug_mode:
            print("Hosts file successfully created")

def copy_hosts_file(remote_server: Server, debug_mode: bool) -> None:
    """ This function is what connects to the actual server in question and runs a ftp command to get the host file onto the machine and then move it into its proper place.
    I had to rewrite this because it's a pain to deal with sudo when using a ssh library. The solution i ended up with was using sftp to just get the file onto the server
    and then using echo "password" | sudo -S cp ..... This is obviously not exactly secure but I really don't care that much since I'm never deploying this code which is famous last words i know.
    The proper way of doing it I guess would be to setup a sshkey for this, but then that kinda gets me back to sqaure one where this is intending to be the first time im connecting to this box so 
    I end up with a chicken and the egg problem with the chicken being the passing the password in plaintext, and the egg being the ssh key which needs sudo to setup.
    I could also probably clear the history of the last 2 commands to make sure I dont leave the password in plaintext, but on my ubuntu server VM i havent seen it so im going to pretend this is 100% fine.

    Args:
        remote_server (Server): The server object which contains the IP and credentials to use
        debug_mode (bool): Whether to print debugging lines
    """
    
    path_of_temp_hosts_file = "hosts"
    remote_path = "/tmp/hosts"

    # Create an SSH client
    ssh = paramiko.SSHClient()
    ssh.set_missing_host_key_policy(paramiko.AutoAddPolicy())
    
    try:
        # Connect to the server
        if debug_mode:
            print(f"Attempting SSH with {remote_server.ip_address} using User:{remote_server._username} & Pass:{remote_server._password}")
        ssh.connect(remote_server.ip_address, username=remote_server._username, password=remote_server._password, timeout=10)

        # Copy the local hosts file to the remote server in a temp dir
        if debug_mode:
            print("Attempting copy")
        sftp = ssh.open_sftp()
        sftp.put(path_of_temp_hosts_file, remote_path)
        if debug_mode:
            print("Host file deployed in /tmp/hosts")
            print("Making backup of hosts file into /tmp/hosts.old")
        # make a backup of the existing remote servers hosts file
        _, stdout, stderr= ssh.exec_command(f"echo {remote_server._password} | sudo -S cp /etc/hosts /tmp/hosts.old")
        if debug_mode:
            print(f"Output: {stdout.read().decode()}")
            print(f"Error: {stderr.read().decode()}")
            print("Attempting move of hosts file from /tmp/hosts to /etc/hosts")
        # move the newly ftped hosts file from /tmp to /etc and override the existing one
        _, stdout, stderr= ssh.exec_command(f"echo {remote_server._password} | sudo -S cp /tmp/hosts /etc/hosts")
        if debug_mode:
            print(f"Output: {stdout.read().decode()}")
            print(f"Error: {stderr.read().decode()}")
        print(f"Hosts file deployed in /etc/hosts for: {remote_server.ip_address} ✅")
    
    except TimeoutError:
        print(f"Connetion timed out for {remote_server.ip_address}\tDid not deploy hosts file ⚠️")
    except OSError:
        print(f"Network is unreachable when attempting to ssh to {remote_server.ip_address}\tDid not deploy hosts file ⚠️")
    except Exception as e:
        print(f"Error: {e}")
    finally:
        ssh.close()


def generate_FQDNs(servers: list[Server], worksheet: worksheet ) -> None:
    """ If everything is where we expect it to be in the Excel worksheet, then we build the fqdns from the domain name corresponding to each subnet

    Args:
        servers (list[Server]): the list of server objects
        worksheet (worksheet): The excel network sheet which would contain the domain name for each subnet
    """
    # make an assumption that they dont have more than 10 different subnets for a single piece of infrastructure which ive never seen
    for column in range(1, 10):
        # This [3,x] coord is by assumption they did not mess up the worksheet 
        if worksheet.cell(3, column).value != None:
            for j in range(len(servers)):
                # check if the cell which should have a subnet name matches the individual servers subnet
                if worksheet.cell(3, column).value == servers[j]._subnet:
                    # check if that subnet has a domain name entry in the cell
                    if worksheet.cell(9, column).value == "None":
                        # If theres no domain name matching the subnet entry 
                        print(f"Could not find a domain name for {servers[j]._subnet}")
                        user_fqdn = input(f"Please enter the domain to use for {servers[j]._hostname}:")
                        servers[j].fqdn = user_fqdn
                    else:
                        servers[j].fqdn = worksheet.cell(9, column).value


def main(servers: list[Server], network_sheet: worksheet, deploy: bool, debug_mode: bool) -> None:
    """ This is the driver for the Host functions

    Args:
        servers (list[Server]): the list of servers
        network_sheet (worksheet): the network sheet of the excel workbook
        deploy (bool): whether we run the copy_hosts_file function
        debug_mode (bool): whether we print the debugging lines
    """
    
    if debug_mode:
        print("\nBefore generate_FQDNS call\n")
        for server in servers:
            print(server)
    # make the host file in the local directory
    generate_FQDNs(servers, network_sheet)
    
    if debug_mode:
        print("\nAfter generate_FQDNS call\n")
        for server in servers:
            print(server)
    
    generate_hosts_file(servers, debug_mode)

    # Copy hosts file to each server
    if deploy:
        for server in servers:
            copy_hosts_file(server, debug_mode)


if __name__ == "__main__":
    # For Testing
    Test_List_of_Servers = [
        Server("DG1234", "192.168.8.3", "3N1D", "subnet1", 10, 20, "username", "p123"),
        Server("DG1234", "192.168.8.4", "3N1D", "subnet1", 10, 20, "u", "p"),
        Server("DG1234", "192.168.8.5", "3N1D", "subnet1", 10, 20, "u", "p"),
        Server("DG1234", "192.168.8.6", "3N1D", "subnet2", 10, 20, "u", "p"),
        Server("DG1234", "192.168.8.7", "3N1D", "subnet2", 10, 20, "u", "p")
    ]
    wb = load_workbook(filename = 'test.xlsx', read_only=True)
    network_sheet = wb["Networks"]
    deploy = False
    debug = True

   

    main(Test_List_of_Servers, network_sheet, deploy, debug)
    