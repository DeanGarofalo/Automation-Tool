"""
Since my code base got very, very messy in C#. I have to put some time into thinking about how this will be organzied.
Do I split each application into folders and classes within it? Sounds dumb
Do I design each helper function as a class and build the driver as a individual app or one big driver?
Or not think at all and start writing code until I see the problem. Likely.
"""

"""
High level functions i need:

☑️ Find_Apps():                     Discover what valid applications are in the workbook, and return those options back to the user to decide what they want to work on, and then launch that apps routine.
☑️ Server class(ip,,pass,etc):      Build the class for the server object for how ill store all the servers information im parsing excel for
☑️ Build_server_object():              Parse the workbook and build out a data structure of the servers found
☑️ Is_valid_config():                  Returns a boolean for if the config is valid or not. If its not valid exit and tell the user to fix the excel


These will be called after the user provides which app they want to work on.
    These will need to take in a int of the number of servers and possibly the type name in a cell
☑️ Find_deployment_type_App-C():     Determine what type of deployment the app is.
    Find_deployment_type_App-B():
    Find_deployment_type_App-D():
    Find_deployment_type_App-S():

Helper functions I need:
☑️ Count_servers():
☑️ Is_valid_IPAddress():
☑️ Is_valid_HA():      Basically look at the count servers output and compare it with what the excel cell says it should be and make sure they match
    Plus way way more

SSH functions I need:
☑️ FTP function
    Firewall function should take in a app and maybe take those ports in as a global var?
    Bash command function
    More I'm forgetting at the moment

TODO ☑️ rewrite support commandline args, should see if i want to do it by hand or use argparse as well. Might be better off because its more versatile    
    
"""

from openpyxl import load_workbook, worksheet, workbook

from Stuff.Apps import C_app
from Stuff import Hosts_helpers, Server

import argparse
import sys
import os

def main():
    # First driver to get the file as input, either through commandline argument on launch or through interactive user input
    parser = argparse.ArgumentParser(prog="Automation App", description="This script takes in a .XLSX file and run sysadmin like functions on its contents.")
    parser.add_argument('filename')
    parser.add_argument('--verbose', '-v', action='store_true', help='Enable verbose mode')
    
    final_filename = ""   
    debug_mode = False

    if len(sys.argv) >= 2:
        args = parser.parse_args()

        # Clean up filename for a few scenarios
        args.filename = sanatize_filepath(args.filename)
        if is_filepath_legit(args.filename):
                final_filename = args.filename
        else:
            while True:
                args.filename = input("Please enter the Excel file path to use: ")
                if is_filepath_legit(args.filename):
                    final_filename = args.filename
                    break
        if args.verbose:
            debug_mode = True
    else:
        print("No arguments provided.")
        while True:
            file_path = input("Please enter the Excel file path to use: ")
            file_path = sanatize_filepath(file_path)
            if is_filepath_legit(file_path):
                final_filename = file_path
                break
    #DEBUG###################################
    if debug_mode:
        print("Full File path:", final_filename)
    #DEBUG###################################
    try:
        #wb = load_workbook(filename = '/mnt/c/Users/dgame/Downloads/Excels/V2/test.xlsx', read_only=True)
        wb = load_workbook(filename = final_filename, read_only=True)
    except:
        print("⛔ That's not a proper Excel file ⛔\nExiting...")
        return
    
    
    # this is where all the Server objects will go
    list_of_servers = []

    # set this down below and use it later for checkspecs, and firewall functions
    what_app_is_this = ""


    selected_app = Find_apps(wb)
    match selected_app:
        case "C" | "C5" | "CB" :
            sheet = wb[selected_app]

            C_app.main(sheet, list_of_servers, debug_mode)
            ...
        case "B":
            sheet = wb[selected_app]
            # run B function
            # B(sheet)
            print("Run B")
            ...
        case "S":
            sheet = wb[selected_app]
            # run S function
            # S(sheet)
            ...
        case _:
            raise ValueError("Invalid Sheet. Exiting")

    #DEBUG#############################
    if debug_mode:
        print("After App.main function")
        for server in list_of_servers:
            print(server)
        print("------------------------------")
    #DEBUG#############################
##########################################################################################################################################################################################
   
   # Now that we have built out the servers and have everything we need at the moment, we prompt the user for what they actually want to do
    while True:
        print("\nWhat would you like to do?")
        print("1) Check Specifications")
        print("2) Make and deploy Hosts file")
        print("3) Open firewall ports")
        print("4) Exit")
        choice = input("\nPlease enter the number corresponding to you choice: ")
        
        match choice:
            case "1": # Check Specs #
                ...
                # checkspecs(list_of_servers, what_app_is_this)
            case "2": # Make & Deploy Hosts file #
                try:
                    network_sheet = wb["Networks"]
                except:
                    print("Could not open Networks tab ⚠️\n")
                    manually_add_FQDNs(list_of_servers)
        
                while True:
                    print("\nDo you also want to deploy the hosts file?")
                    print("1) Yes, deploy the hosts file for me")
                    print("2) No, just generate it")
                    to_deploy_or_not_to_deploy = input("Please enter the number corresponding to you choice: ")
                    if to_deploy_or_not_to_deploy == "1":
                        # deploy
                        Hosts_helpers.main(list_of_servers, network_sheet, True, debug_mode)
                        break
                    if to_deploy_or_not_to_deploy == "2":
                        # not deploy
                        Hosts_helpers.main(list_of_servers, network_sheet, False, debug_mode)
                        print(f"\nHosts file generated in: {os.path.dirname(os.path.realpath(__file__))}/hosts ")
                        break
                    else:
                        print("Invalid input. Please enter 1 or 2 for your choice")
            case "3": # Open firewall ports #
                ...
                # firewallscript(list_of_servers, what_app_is_this)
            case "4": # Exit #
                break
            case _:
                print("Invalid input. Please enter a number for your choice")
        
  
    # End program or add option for advanced mode
    print("END")









def sanatize_filepath(file_path: str) -> str:
    """
    This takes in a str corresponding to a Unix file path location,
    and cleans it up to prevent something unexpected happening.
    We want full file paths to files. So no local/relative path nonsense

    Args:
        file_path (str): str corresponding to a Unix file path location

    Returns:
        str: str corresponding to the full file path location
    """
    if str(file_path).startswith("~"):
        file_path = os.path.expanduser(file_path)
    # Check if its a local or relative path, if local add the full path to the filename
    if not os.path.isabs(file_path):
        file_path = os.path.join(os.getcwd(), file_path)
    return file_path

def is_filepath_legit(file_path: str) -> bool:
    """Check to see that the file exists and is a valid non marco modern Excel file

    Args:
        file_path (str): sanatized str corresponding to the full file path location

    Returns:
        bool: Yes or no answer to the function question
    """
    if os.path.exists(file_path) and os.path.isfile(file_path):
        if not str(file_path).lower().endswith(".xlsx"):
            print("Not an Excel file. Try again")
            return False
        else:
            return True
    else:
        print("Invalid file path or file does not exist. Please enter a valid file path.")
        return False

def Find_apps(workbook: workbook) -> str:
    """
    Print out the discovered sheets and gets the users input for what App they want to work on.

    Args:
        workbook (Openpyxl.Workbook): Excel Workbook 

    Returns:
        selected_sheet (str): Returns the str of the app the user selected which in turn is the sheet to work on.
    """
    # var with supported apps, maybe should be global const instead of here?
    allowed_sheets = ["C", "C5", "D", "B", "S", "Networks", "VoIP"]

    # Get the sheet names
    sheet_names = workbook.sheetnames

    # Print the allowed sheet names with corresponding numbers
    print("Which application would you like to work on?")
    allowed_sheet_indices = []
    for i, name in enumerate(sheet_names, start=1):
        if name in allowed_sheets:
            allowed_sheet_indices.append(i)
            print(f"{len(allowed_sheet_indices)}. {name}")

    # Ask the user for input
    while True:
        try:
            selected_number = int(input("Enter the number of the sheet you want to use: "))
            # If input is in between a valid range
            if 1 <= selected_number <= len(allowed_sheet_indices):
                selected_sheet_index = allowed_sheet_indices[selected_number - 1]
                selected_sheet = sheet_names[selected_sheet_index - 1]
                break
            else:
                print("Invalid number. Please enter a valid number.")
        except ValueError:
            print("Invalid input. Please enter a valid number.")
    return selected_sheet

def manually_add_FQDNs(servers: list[Server.Server] ):
    """Manually ask the user for the fqdn for each server and set it here before needing it in the hosts file creation

    Args:
        servers (list[Server.Server]): The list of servers we are operating on
    """

    print("Entering manual mode for FQDN assignment...\n")

    for server in servers:
        manual_fqdn = input(f"Please enter the domain name to use for {server._hostname}: ")
        server.fqdn = manual_fqdn
        print(f"FQDN set as: {server.fqdn}")
    print("FQDN assignment complete ✅")
    # Could make this ask the user to confirm if they made a mistake and rerun it


###############################################################         EOF          ###########################################################################################################################
if __name__ == "__main__":
    main()